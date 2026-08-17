#!/usr/bin/env python3
"""Переносит фото товаров, встроенные в прайс-листы поставщиков, в каталог.

Что делает:
  1. читает xlsx, достаёт картинки из xl/media и их якоря (какая картинка
     к какой строке привязана);
  2. сопоставляет картинку с артикулом: якорь стоит на первой строке товара,
     строки-варианты идут ниже без названия и наследуют картинку своей группы;
  3. кладёт нормализованные файлы в public/products/sku/;
  4. генерирует src/data/skuPhotos.js — карту «Бренд|Артикул» → файл.

Карта проверяется раньше библиотеки типов (src/lib/productImage.js), потому что
это снимок именно этого товара, а не «такого же типа от того же бренда».

Запуск:  python3 scripts/import-xlsx-photos.py
Требует: openpyxl, ImageMagick (magick).

Прайсы задаются в SOURCES. Проверено: у остальных поставщиков встроенных фото
нет — только логотип на титуле, см. docs/product-images.md.
"""
import json
import os
import re
import subprocess
import zipfile

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO, 'public/products/sku')
MAP_FILE = os.path.join(REPO, 'src/data/skuPhotos.js')

SOURCES = [
    {
        'path': os.path.expanduser('~/Downloads/Menzerna 19.01.2026.xlsx'),
        'sheet': 'Menzerna',
        'brand': 'Menzerna',
        'prefix': 'menzerna',
        'col_sku': 1,     # A — код товара
        'col_name': 2,    # B — наименование, непустое только у первой строки группы
        'col_price': 4,   # D — цена; у строк-заголовков разделов её нет
    },
]


def anchors_by_row(zf):
    """{номер строки: имя файла картинки} по якорям из xl/drawings."""
    out = {}
    for name in zf.namelist():
        m = re.fullmatch(r'xl/drawings/(drawing\d+)\.xml', name)
        if not m:
            continue
        drawing = zf.read(name).decode('utf8')
        rels_path = f'xl/drawings/_rels/{m.group(1)}.xml.rels'
        rels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"',
                               zf.read(rels_path).decode('utf8')))
        for _col, row, rid in re.findall(
                r'<xdr:from><xdr:col>(\d+)</xdr:col>.*?<xdr:row>(\d+)</xdr:row>'
                r'.*?r:embed="(rId\d+)"', drawing, re.S):
            out[int(row) + 1] = rels[rid].split('/')[-1]
    return out


def collect(src):
    """{артикул: имя файла картинки} для одного прайса."""
    wb = openpyxl.load_workbook(src['path'], data_only=True)
    ws = wb[src['sheet']]
    zf = zipfile.ZipFile(src['path'])
    anchors = anchors_by_row(zf)

    sku_img, group_img = {}, None
    for row in range(1, ws.max_row + 1):
        sku = ws.cell(row, src['col_sku']).value
        name = ws.cell(row, src['col_name']).value
        price = ws.cell(row, src['col_price']).value
        if name:                      # началась новая группа товаров
            group_img = anchors.get(row)
        elif row in anchors:          # якорь на строке-варианте
            group_img = anchors[row]
        if sku and price is not None and group_img:
            code = ' '.join(str(sku).split())
            sku_img[code] = group_img
            for part in code.replace('/', ' ').split():   # составные коды
                if len(part) > 4:
                    sku_img.setdefault(part, group_img)
    return zf, sku_img


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    mapping = {}
    for src in SOURCES:
        if not os.path.exists(src['path']):
            print(f"пропущен (нет файла): {src['path']}")
            continue
        zf, sku_img = collect(src)
        files = {}
        for img in sorted(set(sku_img.values())):
            out_name = f"{src['prefix']}-{re.sub(r'[^0-9]', '', img)}.jpg"
            tmp = os.path.join(OUT_DIR, '.tmp-' + img)
            with open(tmp, 'wb') as fh:
                fh.write(zf.read('xl/media/' + img))
            subprocess.run([
                'magick', tmp, '-auto-orient', '-resize', '600x600>',
                '-background', 'white', '-alpha', 'remove', '-alpha', 'off',
                '-gravity', 'center', '-extent', '600x600',
                '-strip', '-quality', '82', os.path.join(OUT_DIR, out_name),
            ], check=True)
            os.remove(tmp)
            files[img] = out_name
        for sku, img in sku_img.items():
            mapping[f"{src['brand']}|{sku}"] = files[img]
        print(f"{os.path.basename(src['path'])}: {len(files)} фото, {len(sku_img)} артикулов")

    lines = [
        '// Фото конкретных артикулов. Приоритетнее общей библиотеки типов:',
        '// это снимок именно этого товара, а не «такого же типа от этого бренда».',
        '//',
        '// Источник — изображения, встроенные в прайс-листы поставщиков.',
        '// Сгенерировано scripts/import-xlsx-photos.py, руками не править.',
        '',
        'export const skuPhotos = {',
    ]
    lines += [f'  {json.dumps(k, ensure_ascii=False)}: {json.dumps(v)},'
              for k, v in sorted(mapping.items())]
    lines += ['};', '']
    with open(MAP_FILE, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(lines))
    print(f'записано {len(mapping)} артикулов -> {os.path.relpath(MAP_FILE, REPO)}')


if __name__ == '__main__':
    main()
