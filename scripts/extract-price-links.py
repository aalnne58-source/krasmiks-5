#!/usr/bin/env python3
"""Достаёт из клиентского прайса ссылки на карточки товаров у поставщика.

В прайсе столбец S «Фото и описание» — это формула
    =ЕСЛИ(A>0; ГИПЕРССЫЛКА(A; "Фото и описание"); "")
то есть сам адрес лежит в столбце A. Ведёт он на HTML-страницу товара,
а не на файл картинки — адрес изображения добирает fetch-supplier-photo-urls.mjs.

Результат: scripts/.cache/price-links.json — [{brand, sku, name, url}].
Файл промежуточный, в git не нужен, пересоздаётся запуском скрипта.

Запуск:  python3 scripts/extract-price-links.py [путь к прайсу.xls]
Требует: openpyxl, LibreOffice (soffice) для чтения .xls.
"""
import json
import os
import subprocess
import sys
import tempfile

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE = os.path.join(REPO, 'scripts/.cache')
OUT = os.path.join(CACHE, 'price-links.json')

DEFAULT_SRC = os.path.expanduser(
    '~/Downloads/прайс_клиента_Шикарная_И_В_ИП_г_Санкт_Петербург_04_06_2026.xls')

COL_URL, COL_BRAND, COL_SKU, COL_NAME = 1, 3, 4, 5
FIRST_ROW = 7  # строки 1-6 — курсы валют и шапка


def to_xlsx(path):
    """openpyxl не читает .xls — конвертируем через LibreOffice."""
    if path.lower().endswith(('.xlsx', '.xlsm')):
        return path, None
    tmp = tempfile.mkdtemp(prefix='price-xls-')
    subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx',
                    '--outdir', tmp, path], check=True, capture_output=True)
    name = os.path.splitext(os.path.basename(path))[0] + '.xlsx'
    return os.path.join(tmp, name), tmp


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    if not os.path.exists(src):
        sys.exit(f'не найден прайс: {src}')

    path, tmp = to_xlsx(src)
    ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]

    rows, no_url = [], 0
    for r in range(FIRST_ROW, ws.max_row + 1):
        brand = ws.cell(r, COL_BRAND).value
        name = ws.cell(r, COL_NAME).value
        if not (brand or name):
            continue
        url = ws.cell(r, COL_URL).value
        url = url.strip() if isinstance(url, str) and url.startswith('http') else None
        if not url:
            no_url += 1
            continue
        rows.append({
            'brand': str(brand or '').strip(),
            'sku': str(ws.cell(r, COL_SKU).value or '').strip(),
            'name': str(name or '').strip(),
            'url': url,
        })

    os.makedirs(CACHE, exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(rows, fh, ensure_ascii=False, indent=1)
    print(f'ссылок: {len(rows)}, строк без ссылки: {no_url}')
    print(f'-> {os.path.relpath(OUT, REPO)}')
    if tmp:
        subprocess.run(['rm', '-rf', tmp], check=False)


if __name__ == '__main__':
    main()
